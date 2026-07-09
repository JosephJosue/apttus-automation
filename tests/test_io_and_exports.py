from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from apttus_delta.config import Config
from apttus_delta.datasets import DATASETS
from apttus_delta.io_excel import IngestError, discover_files, read_dataset, write_workbook
from apttus_delta.soql_inputs import ExportError, check_exports, load_export


def _write_xlsx(path: Path, sheet: str, rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


HEADER = ["Country", "Equip Model", "Equip Desc"]


def test_read_dataset_concats_and_drops_repeated_headers(tmp_path):
    _write_xlsx(tmp_path / "rel/DE/DE_cCRM_Eq_Price_Relevant_List_x.xlsx",
                "tbl_Apttus_cCRM_Eq_Price_Releva",
                [HEADER, ["DE", "M1", "d1"]])
    _write_xlsx(tmp_path / "rel/FR/FR_cCRM_Eq_Price_Relevant_List_x.xlsx",
                "tbl_Apttus_cCRM_Eq_Price_Releva",
                [HEADER, ["FR", "M2", "d2"], ["Country", "Equip Model", "Equip Desc"]])
    files = discover_files(tmp_path / "rel", "cCRM_Eq_Price_Relevant_List")
    assert len(files) == 2
    df = read_dataset(files, "tbl_Apttus_cCRM_Eq_Price_Releva", tuple(HEADER), ("Country",))
    assert df.values.tolist() == [["DE", "M1", "d1"], ["FR", "M2", "d2"]]


def test_read_dataset_rejects_header_drift(tmp_path):
    _write_xlsx(tmp_path / "rel/DE_cCRM_Eq_Price_Relevant_List_x.xlsx",
                "tbl_Apttus_cCRM_Eq_Price_Releva",
                [["Country", "WRONG", "Equip Desc"], ["DE", "M1", "d1"]])
    files = discover_files(tmp_path / "rel", "cCRM_Eq_Price_Relevant_List")
    with pytest.raises(IngestError, match="header mismatch"):
        read_dataset(files, "tbl_Apttus_cCRM_Eq_Price_Releva", tuple(HEADER), ("Country",))


def test_write_workbook_truncates_sheet_names(tmp_path):
    df = pd.DataFrame({"a": ["1"]}, dtype=object)
    out = tmp_path / "o.xlsx"
    write_workbook(out, {"S" * 40: df})
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["S" * 31]


def _cfg(tmp_path, **kw):
    return Config(base_dir=tmp_path, soql_exports=tmp_path / "SOQL Exports",
                  release_date=date(2026, 7, 2), **kw)


def _l01_csv(tmp_path, name, header=None, rows=()):
    folder = tmp_path / "SOQL Exports/prod"
    folder.mkdir(parents=True, exist_ok=True)
    header = header or ["Name", "APTS_Ext_ID__c", "APTS_Country__c",
                        "APTS_Product__r.APTS_Ext_ID__c", "APTS_Product__r.Name"]
    lines = [",".join(header)] + [",".join(r) for r in rows]
    (folder / name).write_text("\n".join(lines), encoding="utf-8")


def test_check_exports_collects_missing_and_stale(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-01-01.csv")  # stale
    found, problems = check_exports(cfg, [DATASETS["L01"], DATASETS["L07"]])
    assert found == []
    assert len(problems) == 2
    assert any("days old" in p for p in problems)
    assert any("no data found" in p for p in problems)


def test_check_exports_picks_newest_and_rejects_ambiguity(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-06-30.csv")
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-07-01.csv")
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert not problems and found[0].export_date == date(2026, 7, 1)

    (tmp_path / "SOQL Exports/prod/Eq_Price_Relevant_List_2026-07-01.xlsx").write_bytes(b"x")
    _, problems = check_exports(cfg, [DATASETS["L01"]])
    assert problems and "ambiguous" in problems[0]


def test_load_export_schema_and_nulls(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-07-01.csv",
             rows=[["M1_DE", "M1_DE", "DE", "M1", ""]])
    found, _ = check_exports(cfg, [DATASETS["L01"]])
    df = load_export(found[0], DATASETS["L01"])
    assert df["APTS_Product__r.Name"].iloc[0] is None  # empty CSV field -> null
    assert df["_"].iloc[0] is None                      # tolerated column added as null


def test_load_export_rejects_schema_drift(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-07-01.csv",
             header=["Name", "APTS_Ext_ID__c", "APTS_Country__c", "Rogue"],
             rows=[["a", "b", "DE", "x"]])
    found, _ = check_exports(cfg, [DATASETS["L01"]])
    with pytest.raises(ExportError, match="schema drift"):
        load_export(found[0], DATASETS["L01"])


def test_config_norway_yaml_boolean(tmp_path):
    cfg_file = tmp_path / "config.yaml"
    cfg_file.write_text("split_countries: [SE, NO, DE]\n", encoding="utf-8")
    from apttus_delta.config import load_config

    cfg = load_config(cfg_file)
    assert cfg.split_countries == ["SE", "NO", "DE"]


def test_config_paths_resolve_under_data_dir(tmp_path):
    data = tmp_path / "somewhere" / "Apttus Automation"
    cfg_file = tmp_path / "cfg/config.yaml"
    cfg_file.parent.mkdir()
    cfg_file.write_text(f'data_dir: "{data.as_posix()}"\n', encoding="utf-8")
    from apttus_delta.config import load_config

    cfg = load_config(cfg_file)
    assert cfg.current_release == data / "01. Current Release"
    assert cfg.soql_exports == data / "SOQL Exports"
    assert cfg.apttus_files == data / "Apttus Files"


def test_config_relative_data_dir_resolves_next_to_config(tmp_path):
    cfg_file = tmp_path / "config.yaml"
    cfg_file.write_text('data_dir: "data"\n', encoding="utf-8")
    from apttus_delta.config import load_config

    cfg = load_config(cfg_file)
    assert cfg.output == tmp_path / "data/Output"


def test_config_cli_data_dir_overrides_file(tmp_path):
    cfg_file = tmp_path / "config.yaml"
    cfg_file.write_text('data_dir: "/ignored"\n', encoding="utf-8")
    from apttus_delta.config import load_config

    cfg = load_config(cfg_file, data_dir=tmp_path / "real")
    assert cfg.previous_release == tmp_path / "real/00. Previous Release"


def _l01_workbook(tmp_path, name, rows=(), header=None, extra_sheets=()):
    folder = tmp_path / "SOQL Exports/prod"
    folder.mkdir(parents=True, exist_ok=True)
    header = header or ["Name", "APTS_Ext_ID__c", "APTS_Country__c",
                        "APTS_Product__r.APTS_Ext_ID__c", "APTS_Product__r.Name"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eq_Price_Relevant_List"
    ws.append(header)
    for r in rows:
        ws.append(r)
    for sheet in extra_sheets:
        wb.create_sheet(title=sheet)
    wb.save(folder / name)


def test_check_exports_uses_workbook_sheet(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_workbook(tmp_path, "SOQL_Exports_2026-07-01.xlsx", rows=[["a", "a", "DE", "M1", "d"]])
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert not problems
    assert found[0].sheet == "Eq_Price_Relevant_List"
    df = load_export(found[0], DATASETS["L01"])
    assert df["APTS_Ext_ID__c"].tolist() == ["a"]


def test_check_exports_empty_workbook_sheet_is_reported(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_workbook(tmp_path, "SOQL_Exports_2026-07-01.xlsx", rows=[])
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert found == []
    assert problems and "has no data" in problems[0]


def test_check_exports_stale_workbook_is_reported(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_workbook(tmp_path, "SOQL_Exports_2026-01-01.xlsx", rows=[["a", "a", "DE", "M1", "d"]])
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert found == []
    assert any("days old" in p for p in problems)


def test_csv_overrides_workbook_when_newer(tmp_path):
    cfg = _cfg(tmp_path)
    _l01_workbook(tmp_path, "SOQL_Exports_2026-06-30.xlsx", rows=[["wb", "wb", "DE", "M", "d"]])
    _l01_csv(tmp_path, "Eq_Price_Relevant_List_2026-07-01.csv",
             rows=[["csv", "csv", "DE", "M", "d"]])
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert not problems and found[0].sheet is None and found[0].path.suffix == ".csv"


def test_workbook_repasted_header_row_is_dropped(tmp_path):
    cfg = _cfg(tmp_path)
    header = ["Name", "APTS_Ext_ID__c", "APTS_Country__c",
              "APTS_Product__r.APTS_Ext_ID__c", "APTS_Product__r.Name"]
    _l01_workbook(tmp_path, "SOQL_Exports_2026-07-01.xlsx",
                  rows=[header, ["a", "a", "DE", "M1", "d"]])
    found, _ = check_exports(cfg, [DATASETS["L01"]])
    df = load_export(found[0], DATASETS["L01"])
    assert df["APTS_Ext_ID__c"].tolist() == ["a"]


def test_make_template_contains_all_soql_sheets(tmp_path):
    from apttus_delta.soql_inputs import write_template

    cfg = _cfg(tmp_path)
    path = write_template(cfg, list(DATASETS.values()))
    wb = openpyxl.load_workbook(path)
    soql_keys = [d.export_key for d in DATASETS.values() if d.mode == "soql"]
    assert wb.sheetnames[0] == "Instructions"
    assert set(soql_keys) <= set(wb.sheetnames)
    ws = wb["Service_Price_Matrix"]
    hdr = [c.value for c in ws[1]]
    assert hdr == list(DATASETS["L07"].org_required)
    # the template itself must not be picked up as a dated export
    found, problems = check_exports(cfg, [DATASETS["L01"]])
    assert found == [] and problems


def test_template_columns_are_text_formatted(tmp_path):
    from apttus_delta.soql_inputs import write_template

    cfg = _cfg(tmp_path)
    path = write_template(cfg, list(DATASETS.values()))
    wb = openpyxl.load_workbook(path)
    ws = wb["CFD_Exhibits"]
    assert ws.column_dimensions["A"].number_format == "@"
    assert ws.column_dimensions["G"].number_format == "@"


def test_config_default_data_dir_is_parent_of_repo(tmp_path):
    """The code lives in python-automation/ inside the data folder, so the
    default data_dir is one level up from where config.yaml sits."""
    repo = tmp_path / "Apttus Automation" / "python-automation"
    repo.mkdir(parents=True)
    cfg_file = repo / "config.yaml"
    cfg_file.write_text('data_dir: ".."\n', encoding="utf-8")
    from apttus_delta.config import load_config

    cfg = load_config(cfg_file)
    assert cfg.current_release.resolve() == (
        tmp_path / "Apttus Automation" / "01. Current Release").resolve()
    assert cfg.soql_exports.resolve() == (
        tmp_path / "Apttus Automation" / "SOQL Exports").resolve()
