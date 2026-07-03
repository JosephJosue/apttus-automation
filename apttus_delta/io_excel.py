"""Release-file discovery, sheet reading and workbook writing.

Reading mirrors the Power Query ingest pattern: find files by a substring
of the file name under a release folder, open the one data sheet, take the
first row as headers, keep every later row (including the repeated header
rows of concatenated country files, which are then dropped by the same
column filters the M code uses)."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

from .normalize import to_text


class IngestError(ValueError):
    pass


def discover_files(root: Path, contains: str, excludes: tuple[str, ...] = ()) -> list[Path]:
    """All .xlsx under `root` (recursive) whose file name contains
    `contains` and none of `excludes` — the SharePoint folder filter of the
    M queries. Sorted for determinism."""
    if not root.is_dir():
        raise IngestError(f"release folder not found: {root}")
    found = []
    for p in sorted(root.rglob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        if contains in p.name and not any(x in p.name for x in excludes):
            found.append(p)
    return found


def _is_lfs_pointer(path: Path) -> bool:
    with open(path, "rb") as f:
        return f.read(30).startswith(b"version https://git-lfs")


def read_sheet_rows(path: Path, sheet: str) -> list[tuple]:
    if _is_lfs_pointer(path):
        raise IngestError(f"{path} is an un-pulled Git LFS pointer; run `git lfs pull` first")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise IngestError(f"{path}: expected sheet {sheet!r}, found {wb.sheetnames}")
        return [tuple(row) for row in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()


def read_dataset(files: list[Path], sheet: str, columns: tuple[str, ...],
                 drop_header_on: tuple[str, ...] = ()) -> pd.DataFrame:
    """Concatenate the data sheet of every file into one all-text frame,
    replicating the M ingest: columns are positional, the first file's
    header row is promoted (and must match `columns`), every later file's
    header row travels along as a data row until the
    `[Country] <> "Country"` style filters drop it. A later file whose
    header text drifts from the expected one gets a warning (Power Query
    would silently swallow it), not an error."""
    if not files:
        raise IngestError(f"no source files found for sheet {sheet!r}")
    n = len(columns)
    data: list[tuple] = []
    for i, path in enumerate(files):
        rows = read_sheet_rows(path, sheet)
        if not rows:
            raise IngestError(f"{path}: sheet {sheet!r} is empty")
        header = tuple(to_text(v) for v in rows[0][:n])
        header += (None,) * (n - len(header))
        if header != tuple(columns):
            if i == 0:
                raise IngestError(
                    f"{path}: header mismatch on sheet {sheet!r}\n"
                    f"  expected: {list(columns)}\n  found:    {list(header)}"
                )
            print(f"warning: {path.name}: header row differs from "
                  f"{files[0].name} on sheet {sheet!r}: {list(header)}", file=sys.stderr)
        body = rows[1:] if i == 0 else rows  # non-first headers stay as data rows
        for row in body:
            vals = tuple(to_text(v) for v in row[:n])
            vals += (None,) * (n - len(vals))
            data.append(vals)
    df = pd.DataFrame(data, columns=list(columns), dtype=object)
    for col in drop_header_on:
        df = df[df[col] != col]
    return df.reset_index(drop=True)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write one sheet per DataFrame, in insertion order, sheet names
    truncated to Excel's 31-character limit."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook(write_only=True)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        ws.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws.append([None if (isinstance(v, float) and v != v) else v for v in row])
    wb.save(path)
