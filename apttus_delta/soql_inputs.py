"""The SOQL export drop-folder contract.

There is no Salesforce API connection: the org side of every SOQL-mode
comparison is supplied manually. The normal way is ONE workbook with one
sheet per query — run each query in Salesforce, copy the result (with
headers) and paste it into the matching sheet, then save the file as:

    SOQL Exports/<org_profile>/SOQL_Exports_<YYYY-MM-DD>.xlsx

where the date is the day the queries were run. `make-template` generates
this workbook pre-filled with the correct headers and an Instructions
sheet holding every query. A per-dataset CSV named
`<export_key>_<YYYY-MM-DD>.csv` is also accepted (it wins over the
workbook for that dataset when newer or equally fresh).

The pipeline picks the newest export per dataset, refuses stale or
ambiguous files, validates the header against the workbook's cCRM query
columns, and applies the same uppercase normalizations that query applies."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

from .datasets import Dataset
from .normalize import to_text

WORKBOOK_PREFIX = "SOQL_Exports"


class ExportError(ValueError):
    pass


@dataclass
class ExportFile:
    dataset_key: str
    export_key: str
    path: Path
    export_date: date
    sheet: str | None = None  # set when the source is a workbook sheet


def _dated_files(folder: Path, stem: str, exts: str = "csv|xlsx"):
    pattern = re.compile(re.escape(stem) + r"_(\d{4}-\d{2}-\d{2})\.(" + exts + r")$")
    matches = []
    if folder.is_dir():
        for p in sorted(folder.iterdir()):
            m = pattern.fullmatch(p.name)
            if m:
                matches.append((date.fromisoformat(m.group(1)), p))
    return matches


def _newest(matches, what: str):
    if not matches:
        return None
    newest = max(d for d, _ in matches)
    files = [p for d, p in matches if d == newest]
    if len(files) > 1:
        raise ExportError(f"{what}: ambiguous exports for {newest} — "
                          + ", ".join(p.name for p in files))
    return ExportFile("", "", files[0], newest)


def _workbook_sheets(path: Path) -> dict[str, int]:
    """Sheet name -> number of data rows (rows after the header)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        counts = {}
        for name in wb.sheetnames:
            n = -1
            for row in wb[name].iter_rows(values_only=True):
                if any(v is not None and v != "" for v in row):
                    n += 1
            counts[name] = max(n, -1)
        return counts
    finally:
        wb.close()


def check_exports(cfg, datasets: list[Dataset]) -> tuple[list[ExportFile], list[str]]:
    """Locate and freshness-check the org export of every SOQL-mode dataset.
    Problems are collected (not fail-on-first) so one run reports every
    missing/stale/empty input."""
    folder = cfg.soql_exports / cfg.org_profile
    found, problems = [], []

    try:
        workbook = _newest(_dated_files(folder, WORKBOOK_PREFIX, "xlsx"), WORKBOOK_PREFIX)
    except ExportError as e:
        return [], [str(e)]
    sheet_rows: dict[str, int] = {}
    if workbook is not None:
        age = (cfg.release_date - workbook.export_date).days
        if age > cfg.export_max_age_days:
            problems.append(
                f"{workbook.path.name} is {age} days old (max {cfg.export_max_age_days}) — "
                f"re-run the queries and save a new dated copy")
            workbook = None
        else:
            sheet_rows = _workbook_sheets(workbook.path)

    for ds in datasets:
        if ds.mode != "soql":
            continue
        try:
            csv_export = _newest(_dated_files(folder, ds.export_key),
                                 f"{ds.key} ({ds.export_key})")
        except ExportError as e:
            problems.append(str(e))
            continue

        wb_has_data = workbook is not None and sheet_rows.get(ds.export_key, -1) > 0
        csv_fresh = (csv_export is not None
                     and (cfg.release_date - csv_export.export_date).days <= cfg.export_max_age_days)

        if csv_fresh and (not wb_has_data or csv_export.export_date >= workbook.export_date):
            found.append(ExportFile(ds.key, ds.export_key, csv_export.path,
                                    csv_export.export_date))
        elif wb_has_data:
            found.append(ExportFile(ds.key, ds.export_key, workbook.path,
                                    workbook.export_date, sheet=ds.export_key))
        elif workbook is not None and ds.export_key in sheet_rows:
            problems.append(
                f"{ds.key} ({ds.label}): sheet '{ds.export_key}' in {workbook.path.name} "
                f"has no data — paste the query result into it")
        elif csv_export is not None:  # exists but stale
            age = (cfg.release_date - csv_export.export_date).days
            problems.append(
                f"{ds.key} ({ds.label}): export {csv_export.path.name} is {age} days old "
                f"(max {cfg.export_max_age_days}) — re-export it")
        else:
            problems.append(
                f"{ds.key} ({ds.label}): no data found — paste the query result into the "
                f"'{ds.export_key}' sheet of {WORKBOOK_PREFIX}_<YYYY-MM-DD>.xlsx in {folder} "
                f"(create it with: python -m apttus_delta make-template)")
    return found, problems


def _read_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, keep_default_na=False, na_values=[],
                     encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL, engine="python")
    return df.astype(object)


def _read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    rows = [r for r in rows if any(v is not None and v != "" for v in r)]
    if not rows:
        raise ExportError(f"{path.name}: sheet {sheet!r} is empty")
    header = [str(to_text(v)).strip() if v is not None else "" for v in rows[0]]
    while header and header[-1] == "":
        header.pop()
    data = []
    for r in rows[1:]:
        vals = [to_text(v) for v in r[: len(header)]]
        vals += [None] * (len(header) - len(vals))
        if all(v == h for v, h in zip(vals, header)):
            continue  # an accidentally re-pasted header row
        data.append(vals)
    return pd.DataFrame(data, columns=header, dtype=object)


def load_export(exp: ExportFile, ds: Dataset, strict_columns: bool = True) -> pd.DataFrame:
    if exp.sheet is not None:
        df = _read_sheet(exp.path, exp.sheet)
        source = f"{exp.path.name} :: {exp.sheet}"
    else:
        df = _read_csv(exp.path) if exp.path.suffix == ".csv" else _read_sheet(
            exp.path, openpyxl.load_workbook(exp.path, read_only=True).sheetnames[0])
        source = exp.path.name
    df.columns = [str(c) for c in df.columns]

    required, optional = set(ds.org_required), set(ds.org_optional)
    have = set(df.columns)
    missing = sorted(required - have)
    extra = sorted(have - required - optional)
    if missing or (extra and strict_columns):
        raise ExportError(
            f"{ds.key} ({ds.label}): schema drift in {source}\n"
            f"  missing columns: {missing or '-'}\n"
            f"  unexpected columns: {extra or '-'}\n"
            f"  expected: {sorted(required)}"
        )
    if extra:
        df = df.drop(columns=extra)
    for col in optional - have:
        df[col] = None

    # Empty cells/fields are nulls, like empty cells in the old pasted table.
    df = df.where(df.notna(), None)
    for col in df.columns:
        df[col] = [None if v == "" else v for v in df[col]]
    for col in ds.org_upper:
        df[col] = df[col].map(lambda v: v.upper() if isinstance(v, str) else v)
    return df


def soql_text(ds: Dataset) -> str:
    """The canonical SELECT for a dataset (same content as docs/soql/*.soql)."""
    fields = ",\n       ".join(ds.org_required)
    where = f"\nWHERE {ds.soql_where}" if ds.soql_where and "=" in ds.soql_where else ""
    note = (f"-- restrict to: {ds.soql_where}\n"
            if ds.soql_where and "=" not in ds.soql_where else "")
    return f"{note}SELECT {fields}\nFROM {ds.soql_object}{where}"


def write_template(cfg, datasets: list[Dataset]) -> Path:
    """One workbook, one pre-headered sheet per SOQL dataset, plus an
    Instructions sheet holding every query. Saved without a date suffix so
    discovery ignores it until the user saves a dated copy."""
    folder = cfg.soql_exports / cfg.org_profile
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{WORKBOOK_PREFIX}_TEMPLATE.xlsx"

    wb = openpyxl.Workbook()
    info = wb.active
    info.title = "Instructions"
    info.append(["How to use this workbook"])
    info.append([f"1. Save a copy named {WORKBOOK_PREFIX}_<date>.xlsx, e.g. "
                 f"{WORKBOOK_PREFIX}_{cfg.release_date.isoformat()}.xlsx (date = today)."])
    info.append(["2. For each sheet: run the query below in Salesforce, copy the full "
                 "result INCLUDING the header row, and paste it into cell A1 of that sheet."])
    info.append(["3. Save, then run: python -m apttus_delta check-exports"])
    info.append([])
    info.append(["Sheet", "Query"])
    for ds in datasets:
        if ds.mode != "soql":
            continue
        info.append([ds.export_key, soql_text(ds)])
        ws = wb.create_sheet(title=ds.export_key)
        ws.append(list(ds.org_required))
        ws.freeze_panes = "A2"
    for cell in info["A"]:
        cell.alignment = openpyxl.styles.Alignment(vertical="top")
    info.column_dimensions["A"].width = 32
    info.column_dimensions["B"].width = 120
    wb.save(path)
    return path
